from selenium import webdriver
import sys
import time
import os
import pyautogui
import webbrowser
import openpyxl
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from openpyxl import Workbook
from openpyxl import load_workbook

chrome_path =r"C:\Users\Desktop\webdriver\chromedriver.exe"
driver = webdriver.Chrome(chrome_path)
driver.get("myurl")

    
def login(driver):

        
    elem = driver.find_element_by_xpath("""//*[@id="usernameField"]""")
    elem.send_keys("myusername")
    elem2 = driver.find_element_by_xpath("""//*[@id="passwordField"]""")
    elem2.send_keys("mypassword")
    driver.find_element_by_xpath("""//*[@id="loginForm"]/table/tbody/tr[4]/td[2]/input""").click()
    driver.find_element_by_xpath("""//*[@id="nav"]/ul/li[2]/a""").click()
    driver.find_element_by_xpath("""//*[@id="check1"]""").click()
    
def sendvalues(driver):
   
    wb = load_workbook('prueba.xlsx')
    ws = wb.active
    buscar = driver.find_element_by_xpath(""" //*[@id="wrapper"]/form/div[2]/input[1]""")
    rut = driver.find_element_by_xpath("""//*[@id="rut"]""")
    dv = driver.find_element_by_xpath("""//*[@id="wrapper"]/form/table[1]/tbody/tr[3]/td[2]/input""")
    nombre = driver.find_element_by_xpath("""//*[@id="wrapper"]/form/table[1]/tbody/tr[4]/td[2]/input""")
    rutvalue=  ws.cell(row=x, column=1).value
    dvvalue= ws.cell(row=x, column=2).value
    nombrevalue= ws.cell(row=x, column=3).value
    rut.send_keys(rutvalue)
    dv.send_keys(dvvalue)
    nombre.send_keys(nombrevalue)
    buscar.click()
    #time.sleep(1)
    table(driver)

def table(driver):
    rut = driver.find_element_by_xpath("""//*[@id="rut"]""")
    dv = driver.find_element_by_xpath("""//*[@id="wrapper"]/form/table[1]/tbody/tr[3]/td[2]/input""")
    nombre = driver.find_element_by_xpath("""//*[@id="wrapper"]/form/table[1]/tbody/tr[4]/td[2]/input""")
    coma = ","
    wb = load_workbook('prueba.xlsx')
    ws = wb.active
    rutvalue=  ws.cell(row=x, column=1).value
    table_elements = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//table[@class = 'grilla']")))
    #time.sleep(1)
    for table_element in table_elements:
        for row in table_element.find_elements_by_xpath(".//tr"):
            text_file = open("Output2.txt", "a")
            #time.sleep(1)
            text_file.write(str(rutvalue)+str(coma)+str(row.text)+'\n')
            #time.sleep(20)
            text_file.close()
            #time.sleep(1)
            rut.clear()
            dv.clear()
            nombre.clear()
                  
    
login(driver)
for x in range(1,5):
   sendvalues(driver)
